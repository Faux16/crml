from __future__ import annotations

import json
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Optional

from ..models.control_catalog_model import CRControlCatalogSchema
from ..models.attack_catalog_model import CRAttackCatalogSchema
from ..models.control_relationships_model import CRControlRelationshipsSchema
from ..models.attack_control_relationships_model import CRAttackControlRelationshipsSchema


_WORKBOOK_FORMAT = "crml_xlsx_mapping"
_WORKBOOK_VERSION = "1.0"

_SHEET_META = "_meta"
_SHEET_CONTROL_CATALOGS = "control_catalogs"
_SHEET_ATTACK_CATALOGS = "attack_catalogs"
_SHEET_CONTROL_RELATIONSHIPS = "control_relationships"
_SHEET_ATTACK_CONTROL_RELATIONSHIPS = "attack_control_relationships"


def _xlsx_module():
    try:
        import openpyxl  # type: ignore

        return openpyxl
    except Exception as e:  # pragma: no cover
        raise ImportError(
            "openpyxl is required for XLSX import/export: pip install 'crml-lang[xlsx]'"
        ) from e


def _to_json_cell(value: Any) -> Optional[str]:
    if value is None:
        return None
    return json.dumps(value, ensure_ascii=False, sort_keys=True)


def _from_json_cell(text: Any) -> Any:
    if text is None:
        return None
    if isinstance(text, str):
        s = text.strip()
        if s == "":
            return None
        return json.loads(s)
    raise TypeError(f"Expected JSON cell as string, got {type(text).__name__}")


def _cell_str(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = str(v)
    return s if s != "" else None


def _cell_float(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        if s == "":
            return None
        return float(s)
    raise TypeError(f"Expected numeric cell, got {type(v).__name__}")


def _now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def _try_int(v: Any) -> Optional[int]:
    if v is None:
        return None
    try:
        return int(v)
    except Exception:
        return None


def _safe_filename(name: str) -> str:
    s = name.strip()
    if not s:
        return "document"

    s = re.sub(r"[^A-Za-z0-9._-]+", "-", s)
    s = re.sub(r"-+", "-", s).strip("-")
    return s or "document"


@dataclass(frozen=True)
class ImportedXlsx:
    control_catalogs: list[CRControlCatalogSchema]
    attack_catalogs: list[CRAttackCatalogSchema]
    control_relationships: list[CRControlRelationshipsSchema]
    attack_control_relationships: list[CRAttackControlRelationshipsSchema]


def export_xlsx(
    out_path: str,
    *,
    control_catalogs: Iterable[CRControlCatalogSchema] = (),
    control_catalog_paths: Iterable[str | Path] = (),
    attack_catalogs: Iterable[CRAttackCatalogSchema] = (),
    attack_catalog_paths: Iterable[str | Path] = (),
    control_relationships: Iterable[CRControlRelationshipsSchema] = (),
    control_relationship_paths: Iterable[str | Path] = (),
    attack_control_relationships: Iterable[CRAttackControlRelationshipsSchema] = (),
    attack_control_relationship_paths: Iterable[str | Path] = (),
) -> None:
    """Export CRML documents into a strict XLSX workbook.

    The workbook is versioned via the `_meta` sheet.
    """

    openpyxl = _xlsx_module()

    catalogs = list(control_catalogs) + _load_docs_from_paths(
        control_catalog_paths, CRControlCatalogSchema
    )
    attack_catalogs_list = list(attack_catalogs) + _load_docs_from_paths(
        attack_catalog_paths, CRAttackCatalogSchema
    )
    rels = list(control_relationships) + _load_docs_from_paths(
        control_relationship_paths, CRControlRelationshipsSchema
    )
    attck_rels = list(attack_control_relationships) + _load_docs_from_paths(
        attack_control_relationship_paths, CRAttackControlRelationshipsSchema
    )

    wb = openpyxl.Workbook()

    # Remove the default sheet.
    if wb.worksheets:
        wb.remove(wb.worksheets[0])

    _write_meta_sheet(wb)
    _write_control_catalogs_sheet(wb, catalogs)
    _write_attack_catalogs_sheet(wb, attack_catalogs_list)
    _write_control_relationships_sheet(wb, rels)
    _write_attack_control_relationships_sheet(wb, attck_rels)

    _apply_workbook_formatting(wb)

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def _apply_workbook_formatting(wb) -> None:
    """Apply purely-presentational formatting to make sheets nicer to use.

    This intentionally does not change the workbook schema.
    """

    openpyxl = _xlsx_module()
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    header_fill = PatternFill("solid", fgColor="F2F2F2")
    header_font = Font(bold=True)
    header_align = Alignment(vertical="top", wrap_text=True)
    wrap_align = Alignment(vertical="top", wrap_text=True)

    def _header_row(ws) -> int:
        # Mapping sheets created by this exporter hide row 1 (machine keys) and
        # show row 2 as the human header.
        return 2 if ws.row_dimensions[1].hidden else 1

    def _style_sheet(ws, *, column_widths: dict[str, float] | None = None) -> int:
        hr = _header_row(ws)
        first_body_row = hr + 1

        # Freeze below the (visible) header row
        ws.freeze_panes = f"A{first_body_row}"

        # Auto-filter across the visible header row
        if ws.max_row >= hr and ws.max_column >= 1:
            ws.auto_filter.ref = f"A{hr}:{get_column_letter(ws.max_column)}{ws.max_row}"

        # Style visible header row
        for cell in ws[hr]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

        # Wrap body cells
        for row in ws.iter_rows(min_row=first_body_row, max_row=ws.max_row, values_only=False):
            for cell in row:
                if cell.value is None:
                    continue
                if isinstance(cell.value, str) and ("{" in cell.value or "[" in cell.value or "\n" in cell.value):
                    cell.alignment = wrap_align

        # Set column widths
        if column_widths:
            for col_name, width in column_widths.items():
                try:
                    idx = _header_index(ws, col_name)
                except KeyError:
                    continue
                ws.column_dimensions[get_column_letter(idx)].width = width

        return first_body_row

    def _add_list_validation(ws, column_name: str, allowed: list[str]) -> None:
        # Excel's list validation uses a comma-separated string in quotes.
        # Keep it short and enum-like.
        try:
            col_idx = _header_index(ws, column_name)
        except KeyError:
            return
        start_row = _header_row(ws) + 1
        if ws.max_row < start_row:
            return

        # Escape double-quotes for Excel list literals by doubling them.
        safe_items = [s.replace('"', '""') for s in allowed]
        formula = '"' + ",".join(safe_items) + '"'
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{get_column_letter(col_idx)}{start_row}:{get_column_letter(col_idx)}{ws.max_row}")

    def _number_format(ws, column_name: str, fmt: str) -> None:
        try:
            col_idx = _header_index(ws, column_name)
        except KeyError:
            return
        start_row = _header_row(ws) + 1
        for r in range(start_row, ws.max_row + 1):
            cell = ws.cell(row=r, column=col_idx)
            if cell.value is None:
                continue
            cell.number_format = fmt

    def _header_index(ws, name: str) -> int:
        header = [c.value for c in ws[1]]
        try:
            return header.index(name) + 1
        except ValueError as e:
            raise KeyError(name) from e

    if _SHEET_META in wb.sheetnames:
        ws = wb[_SHEET_META]
        _style_sheet(ws, column_widths={"format": 22, "version": 12, "created_at": 22, "header_rows": 12})

    if _SHEET_CONTROL_CATALOGS in wb.sheetnames:
        ws = wb[_SHEET_CONTROL_CATALOGS]
        _style_sheet(
            ws,
            column_widths={
                "doc_name": 22,
                "framework": 20,
                "control_id": 18,
                "title": 32,
                "url": 34,
                "tags_json": 22,
                "defense_in_depth_layers_json": 22,
            },
        )
        # Intentionally no Excel data-validation for JSON columns.

    if _SHEET_ATTACK_CATALOGS in wb.sheetnames:
        ws = wb[_SHEET_ATTACK_CATALOGS]
        _style_sheet(
            ws,
            column_widths={
                "doc_name": 22,
                "framework": 26,
                "attack_id": 18,
                "title": 40,
                "url": 34,
                "tags_json": 22,
            },
        )

    if _SHEET_CONTROL_RELATIONSHIPS in wb.sheetnames:
        ws = wb[_SHEET_CONTROL_RELATIONSHIPS]
        _style_sheet(
            ws,
            column_widths={
                "doc_name": 22,
                "source_id": 18,
                "target_id": 18,
                "relationship_type": 16,
                "overlap_weight": 14,
                "overlap_dimensions_json": 26,
                "overlap_rationale": 34,
                "confidence": 12,
                "groupings_json": 26,
                "references_json": 26,
                "description": 34,
            },
        )
        _add_list_validation(
            ws,
            "relationship_type",
            [
                "overlaps_with",
                "mitigates",
                "supports",
                "equivalent_to",
                "parent_of",
                "child_of",
                "backstops",
            ],
        )
        # NOTE: We intentionally do not add Excel data-validation to JSON columns.
        # Excel may remove/strip validations containing nested quotes.
        _number_format(ws, "overlap_weight", "0.00")
        _number_format(ws, "confidence", "0.00")

    if _SHEET_ATTACK_CONTROL_RELATIONSHIPS in wb.sheetnames:
        ws = wb[_SHEET_ATTACK_CONTROL_RELATIONSHIPS]
        _style_sheet(
            ws,
            column_widths={
                "doc_name": 22,
                "attack_id": 18,
                "control_id": 18,
                "relationship_type": 16,
                "strength": 12,
                "confidence": 12,
                "tags_json": 22,
                "references_json": 26,
                "description": 34,
                "metadata_json": 26,
            },
        )
        _add_list_validation(ws, "relationship_type", ["mitigated_by", "detectable_by", "respondable_by"])
        _number_format(ws, "strength", "0.00")
        _number_format(ws, "confidence", "0.00")


def import_xlsx(source: str | Path | Any) -> ImportedXlsx:
    """Import CRML documents from an XLSX workbook created by `export_xlsx`.

    Args:
        source:
            Either a file path (str/Path) or an already-loaded openpyxl Workbook.
    """

    openpyxl = _xlsx_module()

    if isinstance(source, (str, Path)):
        wb = openpyxl.load_workbook(str(source))
    else:
        wb = source
    _validate_meta_sheet(wb)

    header_rows = _get_header_rows(wb)

    return ImportedXlsx(
        control_catalogs=_read_control_catalogs_sheet(wb, header_rows=header_rows),
        attack_catalogs=_read_attack_catalogs_sheet(wb, header_rows=header_rows),
        control_relationships=_read_control_relationships_sheet(wb, header_rows=header_rows),
        attack_control_relationships=_read_attack_control_relationships_sheet(wb, header_rows=header_rows),
    )


def _load_docs_from_paths(paths: Iterable[str | Path], model_cls) -> list[Any]:
    if not paths:
        return []

    from ..yamlio import load_yaml_mapping_from_path

    out = []
    for p in paths:
        data = load_yaml_mapping_from_path(str(p))
        out.append(model_cls.model_validate(data))
    return out


def _write_meta_sheet(wb) -> None:
    ws = wb.create_sheet(_SHEET_META)
    ws.append(["format", "version", "created_at", "header_rows"])
    ws.append([_WORKBOOK_FORMAT, _WORKBOOK_VERSION, _now_iso(), 2])


def _validate_meta_sheet(wb) -> None:
    if _SHEET_META not in wb.sheetnames:
        raise ValueError("XLSX workbook is missing required sheet '_meta'")

    ws = wb[_SHEET_META]
    rows = list(ws.iter_rows(min_row=1, max_row=2, values_only=True))
    if len(rows) < 2:
        raise ValueError("XLSX workbook '_meta' sheet is incomplete")

    header = list(rows[0] or [])
    values = list(rows[1] or [])

    try:
        fmt = values[header.index("format")]
        ver = values[header.index("version")]
    except Exception as e:
        raise ValueError("XLSX workbook '_meta' sheet is malformed") from e

    if fmt != _WORKBOOK_FORMAT:
        raise ValueError(f"Unsupported workbook format: {fmt!r}")
    if ver != _WORKBOOK_VERSION:
        raise ValueError(f"Unsupported workbook version: {ver!r}")


def _get_header_rows(wb) -> int:
    """Number of header rows in mapping sheets.

    Current format always uses 2 header rows:
    - Row 1 (hidden): machine keys
    - Row 2 (visible): human labels
    """

    ws = wb[_SHEET_META]
    rows = list(ws.iter_rows(min_row=1, max_row=2, values_only=True))
    if len(rows) < 2:
        return 1
    header = list(rows[0] or [])
    values = list(rows[1] or [])

    def _get(name: str) -> Any:
        try:
            return values[header.index(name)]
        except Exception:
            return None

    hr = _try_int(_get("header_rows"))
    if hr is None:
        # Treat missing value as current format.
        return 2

    if hr != 2:
        raise ValueError(
            "Unsupported workbook layout: expected _meta.header_rows == 2 for the current format"
        )
    return 2


def _write_human_header(ws, columns: list[tuple[str, str, str]]) -> None:
    """Write a two-row header.

    Row 1: machine keys (hidden)
    Row 2: human labels (visible) + comments with descriptions
    """

    from openpyxl.comments import Comment

    ws.append([c[0] for c in columns])
    ws.append([c[1] for c in columns])

    ws.row_dimensions[1].hidden = True

    for idx, (_, _, desc) in enumerate(columns, start=1):
        if not desc:
            continue
        ws.cell(row=2, column=idx).comment = Comment(desc, "crml")


def _write_control_catalogs_sheet(wb, docs: list[CRControlCatalogSchema]) -> None:
    ws = wb.create_sheet(_SHEET_CONTROL_CATALOGS)
    _write_human_header(
        ws,
        [
            ("doc_name", "Document name", "Name of this CRML document (meta.name)."),
            ("doc_version", "Document version", "Optional document version (meta.version)."),
            ("doc_description", "Document description", "Optional description (meta.description)."),
            ("doc_tags_json", "Document tags (JSON)", "Optional tags array as JSON (e.g. [\"community\"])."),
            ("catalog_id", "Catalog id", "Optional catalog identifier (catalog.id)."),
            ("framework", "Framework", "Framework label (e.g. CIS v8)."),
            ("control_id", "Control id", "Canonical control id (e.g. cisv8:4.2)."),
            ("ref_standard", "Ref standard", "Optional structured ref: standard."),
            ("ref_control", "Ref control", "Optional structured ref: control."),
            ("ref_requirement", "Ref requirement", "Optional structured ref: requirement."),
            ("title", "Title", "Optional short title."),
            ("url", "URL", "Optional URL."),
            ("tags_json", "Tags (JSON)", "Optional tags array as JSON."),
            ("defense_in_depth_layers_json", "Defense-in-depth layers (JSON)", "Optional array as JSON (e.g. [\"prevent\",\"detect\"])."),
        ],
    )

    for doc in docs:
        meta = doc.meta
        cat = doc.catalog
        for entry in cat.controls:
            ref_standard = ref_control = ref_requirement = None
            if entry.ref is not None:
                ref_standard = entry.ref.standard
                ref_control = entry.ref.control
                ref_requirement = entry.ref.requirement

            ws.append(
                [
                    meta.name,
                    meta.version,
                    meta.description,
                    _to_json_cell(meta.tags),
                    cat.id,
                    cat.framework,
                    entry.id,
                    ref_standard,
                    ref_control,
                    ref_requirement,
                    entry.title,
                    entry.url,
                    _to_json_cell(entry.tags),
                    _to_json_cell(entry.defense_in_depth_layers),
                ]
            )


def _write_attack_catalogs_sheet(wb, docs: list[CRAttackCatalogSchema]) -> None:
    ws = wb.create_sheet(_SHEET_ATTACK_CATALOGS)
    _write_human_header(
        ws,
        [
            ("doc_name", "Document name", "Name of this CRML document (meta.name)."),
            ("doc_version", "Document version", "Optional document version (meta.version)."),
            ("doc_description", "Document description", "Optional description (meta.description)."),
            ("doc_tags_json", "Document tags (JSON)", "Optional tags array as JSON."),
            ("catalog_id", "Catalog id", "Optional catalog identifier (catalog.id)."),
            ("framework", "Framework", "Framework label (e.g. MITRE ATT&CK Enterprise)."),
            ("attack_id", "Attack id", "Canonical attack id (e.g. attck:T1059.003)."),
            ("title", "Title", "Optional short title."),
            ("url", "URL", "Optional URL."),
            ("tags_json", "Tags (JSON)", "Optional tags array as JSON."),
            (
                "kill_chain_phases_json",
                "Kill chain phases (JSON)",
                "Optional kill_chain_phases array as JSON (recommended '<kill_chain_name>:<phase_name>').",
            ),
        ],
    )

    for doc in docs:
        meta = doc.meta
        cat = doc.catalog
        for entry in cat.attacks:
            ws.append(
                [
                    meta.name,
                    meta.version,
                    meta.description,
                    _to_json_cell(meta.tags),
                    cat.id,
                    cat.framework,
                    entry.id,
                    entry.title,
                    entry.url,
                    _to_json_cell(entry.tags),
                    _to_json_cell(entry.kill_chain_phases),
                ]
            )


def _read_control_catalogs_sheet(wb, *, header_rows: int) -> list[CRControlCatalogSchema]:
    if _SHEET_CONTROL_CATALOGS not in wb.sheetnames:
        return []

    ws = wb[_SHEET_CONTROL_CATALOGS]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) <= header_rows:
        return []

    header = [str(c) for c in rows[0]]
    out_by_doc: dict[tuple[str, str | None, str | None, str | None, str | None], dict[str, Any]] = {}

    for row in rows[header_rows:]:
        data = dict(zip(header, row))

        doc_name = _cell_str(data.get("doc_name"))
        if not doc_name:
            continue

        doc_version = _cell_str(data.get("doc_version"))
        doc_description = _cell_str(data.get("doc_description"))
        doc_tags = _from_json_cell(data.get("doc_tags_json"))

        catalog_id = _cell_str(data.get("catalog_id"))
        framework = _cell_str(data.get("framework"))
        if not framework:
            raise ValueError(f"control_catalogs row missing framework for doc {doc_name!r}")

        key = (doc_name, doc_version, doc_description, _to_json_cell(doc_tags), catalog_id)
        container = out_by_doc.get(key)
        if container is None:
            container = {
                "crml_control_catalog": "1.0",
                "meta": {
                    "name": doc_name,
                    "version": doc_version,
                    "description": doc_description,
                    "tags": doc_tags,
                },
                "catalog": {
                    "id": catalog_id,
                    "framework": framework,
                    "controls": [],
                },
            }
            out_by_doc[key] = container
        else:
            # framework should be consistent within a doc
            if container["catalog"]["framework"] != framework:
                raise ValueError(
                    f"control_catalogs sheet has conflicting frameworks for doc {doc_name!r}: "
                    f"{container['catalog']['framework']!r} vs {framework!r}"
                )

        ref_standard = _cell_str(data.get("ref_standard"))
        ref_control = _cell_str(data.get("ref_control"))
        ref_requirement = _cell_str(data.get("ref_requirement"))

        ref_obj: Optional[dict[str, Any]] = None
        if ref_standard is not None or ref_control is not None or ref_requirement is not None:
            if ref_standard is None or ref_control is None:
                raise ValueError(
                    f"control_catalogs row has incomplete ref for control {data.get('control_id')!r}"
                )
            ref_obj = {
                "standard": ref_standard,
                "control": ref_control,
                "requirement": ref_requirement,
            }

        container["catalog"]["controls"].append(
            {
                "id": _cell_str(data.get("control_id")),
                "ref": ref_obj,
                "title": _cell_str(data.get("title")),
                "url": _cell_str(data.get("url")),
                "tags": _from_json_cell(data.get("tags_json")),
                "defense_in_depth_layers": _from_json_cell(
                    data.get("defense_in_depth_layers_json")
                ),
            }
        )

    return [CRControlCatalogSchema.model_validate(d) for d in out_by_doc.values()]


def _read_attack_catalogs_sheet(wb, *, header_rows: int) -> list[CRAttackCatalogSchema]:
    if _SHEET_ATTACK_CATALOGS not in wb.sheetnames:
        return []

    ws = wb[_SHEET_ATTACK_CATALOGS]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) <= header_rows:
        return []

    header = [str(c) for c in rows[0]]
    out_by_doc: dict[tuple[str, str | None, str | None, str | None, str | None], dict[str, Any]] = {}

    for row in rows[header_rows:]:
        data = dict(zip(header, row))

        doc_name = _cell_str(data.get("doc_name"))
        if not doc_name:
            continue

        doc_version = _cell_str(data.get("doc_version"))
        doc_description = _cell_str(data.get("doc_description"))
        doc_tags = _from_json_cell(data.get("doc_tags_json"))

        catalog_id = _cell_str(data.get("catalog_id"))
        framework = _cell_str(data.get("framework"))
        if not framework:
            raise ValueError(f"attack_catalogs row missing framework for doc {doc_name!r}")

        key = (doc_name, doc_version, doc_description, _to_json_cell(doc_tags), catalog_id)
        container = out_by_doc.get(key)
        if container is None:
            container = {
                "crml_attack_catalog": "1.0",
                "meta": {
                    "name": doc_name,
                    "version": doc_version,
                    "description": doc_description,
                    "tags": doc_tags,
                },
                "catalog": {
                    "id": catalog_id,
                    "framework": framework,
                    "attacks": [],
                },
            }
            out_by_doc[key] = container
        else:
            if container["catalog"]["framework"] != framework:
                raise ValueError(
                    f"attack_catalogs sheet has conflicting frameworks for doc {doc_name!r}: "
                    f"{container['catalog']['framework']!r} vs {framework!r}"
                )

        container["catalog"]["attacks"].append(
            {
                "id": _cell_str(data.get("attack_id")),
                "title": _cell_str(data.get("title")),
                "url": _cell_str(data.get("url")),
                "tags": _from_json_cell(data.get("tags_json")),
                "kill_chain_phases": _from_json_cell(data.get("kill_chain_phases_json")),
            }
        )

    return [CRAttackCatalogSchema.model_validate(d) for d in out_by_doc.values()]


def _write_control_relationships_sheet(wb, docs: list[CRControlRelationshipsSchema]) -> None:
    ws = wb.create_sheet(_SHEET_CONTROL_RELATIONSHIPS)
    _write_human_header(
        ws,
        [
            ("doc_name", "Document name", "Name of this CRML document (meta.name)."),
            ("doc_version", "Document version", "Optional document version (meta.version)."),
            ("doc_description", "Document description", "Optional description (meta.description)."),
            ("doc_tags_json", "Document tags (JSON)", "Optional tags array as JSON."),
            ("pack_id", "Pack id", "Optional relationship pack identifier (relationships.id)."),
            ("source_id", "Source control id", "Source control (scenario/threat-centric)."),
            ("target_id", "Target control id", "Target control (portfolio/implementation-centric)."),
            (
                "relationship_type",
                "Relationship type",
                "Optional enum: overlaps_with, mitigates, supports, equivalent_to, parent_of, child_of, backstops.",
            ),
            ("overlap_weight", "Overlap weight", "Required overlap weight in [0,1]."),
            ("overlap_dimensions_json", "Overlap dimensions (JSON)", "Optional dimension map as JSON (e.g. {\"coverage\":0.9})."),
            ("overlap_rationale", "Overlap rationale", "Optional rationale/explanation."),
            ("confidence", "Confidence", "Optional confidence in [0,1]."),
            ("groupings_json", "Groupings (JSON)", "Optional groupings array as JSON."),
            ("description", "Description", "Optional free-form description."),
            ("references_json", "References (JSON)", "Optional references array as JSON."),
        ],
    )

    for doc in docs:
        meta = doc.meta
        pack = doc.relationships
        for rel in pack.relationships:
            for target in rel.targets:
                ws.append(
                    [
                        meta.name,
                        meta.version,
                        meta.description,
                        _to_json_cell(meta.tags),
                        pack.id,
                        rel.source,
                        target.target,
                        target.relationship_type,
                        target.overlap.weight,
                        _to_json_cell(target.overlap.dimensions),
                        target.overlap.rationale,
                        target.confidence,
                        _to_json_cell(
                            [g.model_dump(exclude_none=True) for g in (target.groupings or [])]
                        )
                        if target.groupings
                        else None,
                        target.description,
                        _to_json_cell(
                            [r.model_dump(exclude_none=True) for r in (target.references or [])]
                        )
                        if target.references
                        else None,
                    ]
                )


def _read_control_relationships_sheet(
    wb, *, header_rows: int
) -> list[CRControlRelationshipsSchema]:
    if _SHEET_CONTROL_RELATIONSHIPS not in wb.sheetnames:
        return []

    ws = wb[_SHEET_CONTROL_RELATIONSHIPS]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) <= header_rows:
        return []

    header = [str(c) for c in rows[0]]
    out_by_doc: dict[tuple[str, str | None, str | None, str | None, str | None], dict[str, Any]] = {}

    def _doc_key(doc_name: str, doc_version, doc_description, doc_tags, pack_id):
        return (doc_name, doc_version, doc_description, _to_json_cell(doc_tags), pack_id)

    for row in rows[header_rows:]:
        data = dict(zip(header, row))
        doc_name = _cell_str(data.get("doc_name"))
        if not doc_name:
            continue

        doc_version = _cell_str(data.get("doc_version"))
        doc_description = _cell_str(data.get("doc_description"))
        doc_tags = _from_json_cell(data.get("doc_tags_json"))
        pack_id = _cell_str(data.get("pack_id"))

        key = _doc_key(doc_name, doc_version, doc_description, doc_tags, pack_id)
        container = out_by_doc.get(key)
        if container is None:
            container = {
                "crml_control_relationships": "1.0",
                "meta": {
                    "name": doc_name,
                    "version": doc_version,
                    "description": doc_description,
                    "tags": doc_tags,
                },
                "relationships": {
                    "id": pack_id,
                    "relationships": [],
                },
            }
            out_by_doc[key] = container

        source_id = _cell_str(data.get("source_id"))
        target_id = _cell_str(data.get("target_id"))
        if not source_id or not target_id:
            raise ValueError(f"control_relationships row missing source/target in doc {doc_name!r}")

        overlap_weight = _cell_float(data.get("overlap_weight"))
        if overlap_weight is None:
            raise ValueError(
                f"control_relationships row missing overlap_weight for {source_id!r} -> {target_id!r}"
            )

        relationship_type = _cell_str(data.get("relationship_type"))
        overlap_dimensions = _from_json_cell(data.get("overlap_dimensions_json"))
        overlap_rationale = _cell_str(data.get("overlap_rationale"))
        confidence = _cell_float(data.get("confidence"))
        groupings = _from_json_cell(data.get("groupings_json"))
        description = _cell_str(data.get("description"))
        references = _from_json_cell(data.get("references_json"))

        # Group by source, preserve row order for targets.
        rel_list = container["relationships"]["relationships"]
        if rel_list and rel_list[-1]["source"] == source_id:
            grouped = rel_list[-1]
        else:
            grouped = {"source": source_id, "targets": []}
            rel_list.append(grouped)

        grouped["targets"].append(
            {
                "target": target_id,
                "relationship_type": relationship_type,
                "overlap": {
                    "weight": overlap_weight,
                    "dimensions": overlap_dimensions,
                    "rationale": overlap_rationale,
                },
                "confidence": confidence,
                "groupings": groupings,
                "description": description,
                "references": references,
            }
        )

    return [CRControlRelationshipsSchema.model_validate(d) for d in out_by_doc.values()]


def _write_attack_control_relationships_sheet(
    wb, docs: list[CRAttackControlRelationshipsSchema]
) -> None:
    ws = wb.create_sheet(_SHEET_ATTACK_CONTROL_RELATIONSHIPS)
    _write_human_header(
        ws,
        [
            ("doc_name", "Document name", "Name of this CRML document (meta.name)."),
            ("doc_version", "Document version", "Optional document version (meta.version)."),
            ("doc_description", "Document description", "Optional description (meta.description)."),
            ("doc_tags_json", "Document tags (JSON)", "Optional tags array as JSON."),
            ("pack_id", "Pack id", "Optional relationship pack identifier (relationships.id)."),
            ("attack_id", "Attack id", "Attack pattern id (e.g. attck:T1059.003)."),
            ("control_id", "Control id", "Mapped control id (e.g. cap:edr)."),
            (
                "relationship_type",
                "Relationship type",
                "Enum: mitigated_by, detectable_by, respondable_by.",
            ),
            ("strength", "Strength", "Optional strength in [0,1]."),
            ("confidence", "Confidence", "Optional confidence in [0,1]."),
            ("description", "Description", "Optional free-form description."),
            ("tags_json", "Tags (JSON)", "Optional tags array as JSON."),
            ("references_json", "References (JSON)", "Optional references array as JSON."),
            ("metadata_json", "Pack metadata (JSON)", "Optional pack-level metadata map as JSON."),
        ],
    )

    for doc in docs:
        meta = doc.meta
        pack = doc.relationships
        for rel in pack.relationships:
            for tgt in rel.targets:
                ws.append(
                    [
                        meta.name,
                        meta.version,
                        meta.description,
                        _to_json_cell(meta.tags),
                        pack.id,
                        rel.attack,
                        tgt.control,
                        tgt.relationship_type,
                        tgt.strength,
                        tgt.confidence,
                        tgt.description,
                        _to_json_cell(tgt.tags),
                        _to_json_cell(
                            [r.model_dump(exclude_none=True) for r in (tgt.references or [])]
                        )
                        if tgt.references
                        else None,
                        _to_json_cell(pack.metadata),
                    ]
                )


def _read_attack_control_relationships_sheet(
    wb, *, header_rows: int
) -> list[CRAttackControlRelationshipsSchema]:
    if _SHEET_ATTACK_CONTROL_RELATIONSHIPS not in wb.sheetnames:
        return []

    ws = wb[_SHEET_ATTACK_CONTROL_RELATIONSHIPS]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) <= header_rows:
        return []

    header = [str(c) for c in rows[0]]
    out_by_doc: dict[tuple[str, str | None, str | None, str | None, str | None], dict[str, Any]] = {}

    def _doc_key(doc_name: str, doc_version, doc_description, doc_tags, pack_id):
        return (doc_name, doc_version, doc_description, _to_json_cell(doc_tags), pack_id)

    for row in rows[header_rows:]:
        data = dict(zip(header, row))
        doc_name = _cell_str(data.get("doc_name"))
        if not doc_name:
            continue

        doc_version = _cell_str(data.get("doc_version"))
        doc_description = _cell_str(data.get("doc_description"))
        doc_tags = _from_json_cell(data.get("doc_tags_json"))
        pack_id = _cell_str(data.get("pack_id"))
        metadata = _from_json_cell(data.get("metadata_json"))

        key = _doc_key(doc_name, doc_version, doc_description, doc_tags, pack_id)
        container = out_by_doc.get(key)
        if container is None:
            container = {
                "crml_attack_control_relationships": "1.0",
                "meta": {
                    "name": doc_name,
                    "version": doc_version,
                    "description": doc_description,
                    "tags": doc_tags,
                },
                "relationships": {
                    "id": pack_id,
                    "relationships": [],
                    "metadata": metadata,
                },
            }
            out_by_doc[key] = container
        else:
            if container["relationships"].get("metadata") != metadata:
                raise ValueError(
                    f"attack_control_relationships sheet has conflicting metadata for doc {doc_name!r}"
                )

        attack_id = _cell_str(data.get("attack_id"))
        control_id = _cell_str(data.get("control_id"))
        relationship_type = _cell_str(data.get("relationship_type"))
        if not attack_id or not control_id or not relationship_type:
            raise ValueError(
                f"attack_control_relationships row missing required fields in doc {doc_name!r}"
            )

        strength = _cell_float(data.get("strength"))
        confidence = _cell_float(data.get("confidence"))
        description = _cell_str(data.get("description"))
        tags = _from_json_cell(data.get("tags_json"))
        references = _from_json_cell(data.get("references_json"))

        rel_list = container["relationships"]["relationships"]
        if rel_list and rel_list[-1]["attack"] == attack_id:
            grouped = rel_list[-1]
        else:
            grouped = {"attack": attack_id, "targets": []}
            rel_list.append(grouped)

        grouped["targets"].append(
            {
                "control": control_id,
                "relationship_type": relationship_type,
                "strength": strength,
                "confidence": confidence,
                "description": description,
                "references": references,
                "tags": tags,
            }
        )

    return [CRAttackControlRelationshipsSchema.model_validate(d) for d in out_by_doc.values()]


def write_imported_as_yaml(
    imported: ImportedXlsx, out_dir: str, *, overwrite: bool = False, sort_keys: bool = False
) -> list[str]:
    """Write imported documents to YAML files.

    Returns the list of file paths written.
    """

    from ..yamlio import dump_yaml_to_path

    out = []
    Path(out_dir).mkdir(parents=True, exist_ok=True)

    def _write(doc_type: str, doc_name: str, payload: dict[str, Any]) -> str:
        base = _safe_filename(doc_name)
        path = str(Path(out_dir) / f"{base}-{doc_type}.yaml")
        if not overwrite and Path(path).exists():
            raise FileExistsError(f"Refusing to overwrite existing file: {path}")
        dump_yaml_to_path(payload, path, sort_keys=sort_keys)
        return path

    for doc in imported.control_catalogs:
        out.append(_write("control-catalog", doc.meta.name, doc.model_dump(by_alias=True, exclude_none=True)))
    for doc in imported.attack_catalogs:
        out.append(_write("attack-catalog", doc.meta.name, doc.model_dump(by_alias=True, exclude_none=True)))
    for doc in imported.control_relationships:
        out.append(
            _write(
                "control-relationships",
                doc.meta.name,
                doc.model_dump(by_alias=True, exclude_none=True),
            )
        )
    for doc in imported.attack_control_relationships:
        out.append(
            _write(
                "attack-control-relationships",
                doc.meta.name,
                doc.model_dump(by_alias=True, exclude_none=True),
            )
        )

    return out
