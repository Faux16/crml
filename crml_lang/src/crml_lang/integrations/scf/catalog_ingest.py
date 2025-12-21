import logging
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

from crml_lang.models.control_catalog_model import (
    ControlCatalogEntry,
    ControlCatalog,
    CRControlCatalog,
)
from crml_lang.models.scenario_model import Meta
from crml_lang.integrations.scf._require import require_scf

logger = logging.getLogger(__name__)

# Likely column names in SCF Excel
COL_SCF_ID = ["SCF #", "SCF Control", "Control ID"]
COL_DESCRIPTION = ["Control Description", "Description"]
COL_QUESTION = ["Control Question", "Question"]
COL_DOMAIN = ["Domain", "Family"]
COL_CAPABILITY = ["Capability"]

def _find_column_index(headers: List[str], candidates: List[str]) -> int:
    """Return the index of the first matching candidate header, or -1."""
    if not headers:
        return -1
    lower_headers = [h.lower().strip() if h else "" for h in headers]
    for c in candidates:
        if c.lower() in lower_headers:
            return lower_headers.index(c.lower())
    return -1

def read_scf_catalog_as_crml(path: Union[str, Path]) -> CRControlCatalog:
    """
    Reads an SCF Excel file and converts it to a CRML ControlCatalog.
    """
    require_scf()
    import openpyxl

    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"SCF file not found: {path}")

    wb = openpyxl.load_workbook(filename=path, read_only=True, data_only=True)
    
    # Try to find the relevant sheet. Usually "SCF 20xx" or just the active one.
    sheet = wb.active
    # If there's a sheet with "SCF" in the name, prefer that.
    for sheet_name in wb.sheetnames:
        if "SCF" in sheet_name and "20" in sheet_name: # Matches "SCF 2024.1" etc.
            sheet = wb[sheet_name]
            break

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("SCF file seems empty.")

    # Find header row (usually not the first one, might be row 2 or 3)
    header_idx = -1
    col_map = {}

    for i, row in enumerate(rows[:20]): # Scan first 20 rows
        str_row = [str(c) if c is not None else "" for c in row]
        idx_id = _find_column_index(str_row, COL_SCF_ID)
        
        if idx_id != -1:
            header_idx = i
            col_map['id'] = idx_id
            col_map['domain'] = _find_column_index(str_row, COL_DOMAIN)
            col_map['question'] = _find_column_index(str_row, COL_QUESTION)
            col_map['capability'] = _find_column_index(str_row, COL_CAPABILITY)
            break
            
    if header_idx == -1:
        raise ValueError("Could not identify SCF headers. Looked for 'SCF #'")

    controls: List[ControlCatalogEntry] = []

    # Iterate data rows
    for row in rows[header_idx+1:]:
        if not row:
            continue
            
        def get_val(key):
            idx = col_map.get(key, -1)
            if idx != -1 and idx < len(row):
                val = row[idx]
                return str(val).strip() if val is not None else None
            return None

        c_id = get_val('id')
        if not c_id:
            continue
            
        # Ensure it has a namespace. CRML demands namespaced IDs.
        # Also remove spaces from the ID part
        c_id = c_id.replace(" ", "")
        if ":" not in c_id:
            c_id = f"scf:{c_id}"

        c_question = get_val('question')
        c_domain = get_val('domain')
        
        # Use Question as title if available
        title = c_question if c_question else f"SCF Control {c_id}"
        
        tags = []
        if c_domain:
            tags.append(c_domain)
            
        entry = ControlCatalogEntry(
            id=c_id,
            title=title,
            tags=tags if tags else None,
        )
        controls.append(entry)

    wb.close()
    
    # Construct Catalog
    catalog_payload = ControlCatalog(
        framework="Secure Controls Framework (SCF)",
        controls=controls
    )

    return CRControlCatalog(
        crml_control_catalog="1.0",
        meta=Meta(
            name="Imported SCF Catalog",
            description=f"Imported from {path.name}",
            version="1.0"
        ),
        catalog=catalog_payload
    )
