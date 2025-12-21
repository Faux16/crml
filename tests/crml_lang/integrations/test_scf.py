import pytest
from openpyxl import Workbook
from crml_lang.integrations.scf import read_scf_catalog_as_crml

@pytest.fixture
def scf_excel_file(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["SCF #", "Control Question", "Domain", "Control Description", "Privacy", "Security"])
    ws.append(["AC-1", "Does the org have access control?", "Access Control", "Details here", "Yes", "Yes"])
    ws.append(["Net-1", "Is network secure?", "Network Security", "More details", "No", "Yes"])
    
    path = tmp_path / "scf_test.xlsx"
    wb.save(path)
    return str(path)

def test_read_scf_catalog_basic(scf_excel_file):
    """Test standard SCF catalog import."""
    catalog = read_scf_catalog_as_crml(scf_excel_file)
    assert catalog.crml_control_catalog == "1.0"
    assert len(catalog.catalog.controls) == 2
    
    c1 = catalog.catalog.controls[0]
    assert c1.id == "scf:AC-1"
    assert c1.title == "Does the org have access control?"
    assert c1.tags == ["Access Control"]
    
    c2 = catalog.catalog.controls[1]
    assert c2.id == "scf:Net-1"
    assert c2.title == "Is network secure?"
    assert c2.tags == ["Network Security"]

def test_scf_id_normalization(tmp_path):
    """Test that IDs are normalized (spaces removed, namespace added)."""
    wb = Workbook()
    ws = wb.active
    ws.append(["SCF #", "Control Question", "Domain"])
    ws.append([" AC-1 ", "Q1", "D1"]) # Spaces
    ws.append(["test:ID-2", "Q2", "D1"]) # Already namespaced
    
    path = tmp_path / "scf_norm.xlsx"
    wb.save(path)
    
    catalog = read_scf_catalog_as_crml(str(path))
    assert len(catalog.catalog.controls) == 2
    assert catalog.catalog.controls[0].id == "scf:AC-1"
    
    # My logic preserves existing namespaces if ":" is present
    assert catalog.catalog.controls[1].id == "test:ID-2"

def test_missing_required_columns(tmp_path):
    """Test that missing required columns raises an error."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Wrong Column", "Another"])
    path = tmp_path / "bad.xlsx"
    wb.save(path)
    
    with pytest.raises(ValueError, match="Could not identify SCF headers"):
        read_scf_catalog_as_crml(str(path))

def test_empty_rows(tmp_path):
    """Test that empty rows are skipped."""
    wb = Workbook()
    ws = wb.active
    ws.append(["SCF #", "Control Question", "Domain"])
    ws.append(["AC-1", "Q1", "D1"])
    ws.append([None, None, None]) # Empty
    ws.append(["AC-2", "Q2", "D2"])
    
    path = tmp_path / "empty_rows.xlsx"
    wb.save(path)
    
    catalog = read_scf_catalog_as_crml(str(path))
    assert len(catalog.catalog.controls) == 2
    assert catalog.catalog.controls[0].id == "scf:AC-1"
    assert catalog.catalog.controls[1].id == "scf:AC-2"
